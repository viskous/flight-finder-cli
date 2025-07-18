from tabulate import tabulate
from colorama import Fore, Style
import requests, shutil, os
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# loading the env file containing api key
load_dotenv()

# getting the terminal width
columns = shutil.get_terminal_size().columns

# fetching credentials from environment variables
API_KEY = os.getenv("AMADEUS_API_KEY")
API_SECRET = os.getenv("AMADEUS_API_SECRET")

# clearing the terminal just to keep things clean
def clear_terminal():
    os.system('cls' if os.name == 'nt' else 'clear')

# gets the access token from amadeus using client credentials
def get_access_token():
    url = "https://test.api.amadeus.com/v1/security/oauth2/token"
    payload = {
        "grant_type": "client_credentials",
        "client_id": API_KEY,
        "client_secret": API_SECRET
    }
    response = requests.post(url, data=payload)
    return response.json().get("access_token")

# gets airline name using carrier code (uses cache to avoid duplicate requests)
def get_airline_name(code, token, cache={}):
    if code in cache:
        return cache[code]

    url = f"https://test.api.amadeus.com/v1/reference-data/airlines?airlineCodes={code}"
    headers = {"Authorization": f"Bearer {token}"}
    response = requests.get(url, headers=headers)

    if response.status_code == 200:
        data = response.json().get("data", [])
        if data:
            name = data[0].get("businessName") or data[0].get("commonName") or data[0].get("name")
            cache[code] = name
            return name

    # fallback to code itself if no name found
    cache[code] = code
    return code

# handles errors when API call to search flights fails
def _handle_search_error(response):
    try:
        error = response.json()
        problem = error.get('errors', [{}])[0].get('title', 'unknown error')
        detail = error.get('errors', [{}])[0].get('detail', 'no additional details')
    except Exception:
        problem = "unexpected error"
        detail = f"http status code: {response.status_code}"

    print(Fore.RED + f"""
❌ Error fetching flights:
        Problem: {problem}
        Detail : {detail}
""" + Style.RESET_ALL)

# uses amadeus flight search API to get flight data
def search_flights(origin, destination, date, travel_class):
    token = get_access_token()
    headers = {"Authorization": f"Bearer {token}"}
    params = {
        "originLocationCode": origin,
        "destinationLocationCode": destination,
        "departureDate": date,
        "adults": 1,
        "max": 5,
        "currencyCode": "USD",
        "travelClass": travel_class.upper()
    }

    url = "https://test.api.amadeus.com/v2/shopping/flight-offers"
    response = requests.get(url, headers=headers, params=params)

    if response.status_code != 200:
        _handle_search_error(response)
        return [], None

    return response.json().get("data", []), token

# prepares a dictionary for each flight with cleaned and formatted values
def _prepare_flight_dict(flight, token):
    price = float(flight['price']['total'])
    segments = flight['itineraries'][0]['segments']
    stops = len(segments) - 1

    dep = segments[0]['departure']
    arr = segments[-1]['arrival']

    airline_code = segments[0]['carrierCode']
    flight_number = segments[0].get('number', '')
    airline_name = get_airline_name(airline_code, token)
    airline = f"{airline_name} {flight_number}"

    dep_code = dep['iataCode']
    arr_code = arr['iataCode']
    dep_date, dep_time = dep['at'].split('T')
    arr_date, arr_time = arr['at'].split('T')
    duration = flight['itineraries'][0]['duration'].replace("PT", "").lower()

    # get all stop airport codes if multiple segments
    stop_airports = (
        " → ".join([seg['arrival']['iataCode'] for seg in segments[:-1]])
        if stops > 0 else "Non-stop"
    )

    return {
        "Airline": airline,
        "DEP": dep_code,
        "Departure Date": dep_date,
        "Departure Time": dep_time,
        "ARR": arr_code,
        "Arrival Date": arr_date,
        "Arrival Time": arr_time,
        "Duration": duration,
        "Stops": f"{stops} ({stop_airports})",
        "Price (USD)": price
    }

# sorts the flight list based on selected field (price / duration / departure time)
def _sort_flight_list(flights, sort_by):
    if sort_by == "price":
        flights.sort(key=lambda x: x["Price (USD)"])
    elif sort_by == "departure":
        flights.sort(key=lambda x: f"{x['Departure Date']} {x['Departure Time']}")
    elif sort_by == "duration":
        # converts duration like "2h30m" to total minutes
        def duration_to_minutes(dur):
            h, m = 0, 0
            if 'h' in dur:
                parts = dur.split('h')
                h = int(parts[0])
                m = int(parts[1].replace('m', '')) if 'm' in parts[1] else 0
            elif 'm' in dur:
                m = int(dur.replace('m', ''))
            return h * 60 + m
        flights.sort(key=lambda x: duration_to_minutes(x["Duration"]))

# formats the header row of the Excel sheet (bold + centered)
def _style_excel_header(ws, headers):
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

# writes the flight data into Excel rows
def _write_excel_rows(ws, display_list, headers):
    for row_index, flight in enumerate(display_list, start=2):
        for col_index, key in enumerate(headers, start=1):
            cell = ws.cell(row=row_index, column=col_index)
            cell.value = flight[key]
            cell.alignment = Alignment(horizontal="center")

# main function to write sorted flight data to Excel file
def write_flights_to_excel(flights, sort_by, origin, destination, date, travel_class, token):
    if not flights:
        print("🚫 No flights found.")
        return

    display_list = [_prepare_flight_dict(flight, token) for flight in flights]
    _sort_flight_list(display_list, sort_by)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Flight Results"

    headers = [
        "Airline", "DEP", "Departure Date", "Departure Time", "ARR",
        "Arrival Date", "Arrival Time", "Duration", "Stops", "Price (USD)"
    ]

    ws.append(headers)
    _style_excel_header(ws, headers)
    _write_excel_rows(ws, display_list, headers)

    # auto-adjust column width for readability
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    wb.save("flights_result.xlsx")
    print("✅ Flights saved to 'flights_result.xlsx'".center(columns - 1))

# driver function that runs the CLI interface and calls everything
def main():
    clear_terminal()

    # header banner
    print(Fore.BLUE + "-" * columns + Style.RESET_ALL)
    print(Fore.CYAN + "✈️  FLIGHT FINDER".center(columns) + Style.RESET_ALL)
    print(Fore.BLUE + "-" * columns + Style.RESET_ALL)

    # getting user inputs
    origin       = input("🔹 From (IATA code, e.g., DEL)                 : ").strip().upper()
    destination  = input("🔹 To (IATA code, e.g., BOM)                   : ").strip().upper()
    date         = input("📅 Departure Date (YYYY-MM-DD)                 : ").strip()
    travel_class = input("💺 Class (Economy / Premium / Business / First): ").strip().lower()
    sort_by      = input("🔽 Sort by (price / departure / duration)      : ").strip().lower()

    print("")

    # search flights and save if available
    flights, token = search_flights(origin, destination, date, travel_class)

    if flights:
        write_flights_to_excel(flights, sort_by, origin, destination, date, travel_class, token)

    print("")
    print(Fore.BLUE + "-" * columns + Style.RESET_ALL)
    print(Fore.CYAN + "✅ Search Complete".center(columns - 1) + Style.RESET_ALL)
    print(Fore.BLUE + "-" * columns + Style.RESET_ALL)

# main guard to run script
if __name__ == "__main__":
    main()
